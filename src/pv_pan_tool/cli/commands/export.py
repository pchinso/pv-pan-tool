"""
Export command for PV PAN Tool CLI.

This module provides functionality to export module data
to various formats (CSV, Excel, JSON).
"""

from pathlib import Path

import click
from rich.console import Console
from rich.panel import Panel
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
)

from ...database import PVModuleDatabase
from ..utils.config import get_config
from ..utils.formatters import format_csv, format_json

console = Console()


@click.command()
@click.option(
    "--format", "-f",
    "export_format",
    type=click.Choice(['csv', 'xlsx', 'json']),
    default='csv',
    help="Export format"
)
@click.option(
    "--output", "-o",
    type=click.Path(path_type=Path),
    required=True,
    help="Output file path"
)
@click.option(
    "--manufacturer", "-m",
    help="Filter by manufacturer name (partial match)"
)
@click.option(
    "--model", "-M",
    help="Filter by model name (partial match)"
)
@click.option(
    "--power-min", "-p",
    type=float,
    help="Minimum power rating (W)"
)
@click.option(
    "--power-max", "-P",
    type=float,
    help="Maximum power rating (W)"
)
@click.option(
    "--efficiency-min", "-e",
    type=float,
    help="Minimum efficiency (%)"
)
@click.option(
    "--efficiency-max", "-E",
    type=float,
    help="Maximum efficiency (%)"
)
@click.option(
    "--cell-type", "-c",
    type=click.Choice(['monocrystalline', 'polycrystalline', 'thin_film', 'perc', 'bifacial', 'hjt', 'ibc']),
    help="Filter by cell type"
)
@click.option(
    "--module-type", "-t",
    type=click.Choice(['standard', 'bifacial', 'glass_glass', 'flexible', 'bipv']),
    help="Filter by module type"
)
@click.option(
    "--height-min", "-h",
    type=float,
    help="Minimum height (mm)"
)
@click.option(
    "--height-max", "-H",
    type=float,
    help="Maximum height (mm)"
)
@click.option(
    "--width-min", "-w",
    type=float,
    help="Minimum width (mm)"
)
@click.option(
    "--width-max", "-W",
    type=float,
    help="Maximum width (mm)"
)
@click.option(
    "--include-metadata", "-i",
    is_flag=True,
    help="Include file metadata in export"
)
@click.option(
    "--include-raw", "-r",
    is_flag=True,
    help="Include raw .PAN file data"
)
@click.option(
    "--sort-by", "-s",
    type=click.Choice(['pmax_stc', 'efficiency_stc', 'manufacturer', 'model', 'voc_stc', 'isc_stc']),
    default='pmax_stc',
    help="Sort results by field"
)
@click.option(
    "--sort-order", "-S",
    type=click.Choice(['asc', 'desc']),
    default='desc',
    help="Sort order"
)
@click.option(
    "--limit", "-l",
    type=int,
    help="Maximum number of records to export"
)
@click.pass_context
def export(ctx, export_format, output, manufacturer, model, power_min, power_max,
          efficiency_min, efficiency_max, cell_type, module_type,
          height_min, height_max, width_min, width_max,
          include_metadata, include_raw, sort_by, sort_order, limit):
    """
    Export module data to various formats.

    Export PV module data from the database to CSV, Excel, or JSON formats
    with optional filtering and sorting.

    Examples:
        pv-pan-tool export --format csv --output modules.csv
        pv-pan-tool export --format xlsx --manufacturer "Jinko" --output jinko.xlsx
        pv-pan-tool export --format json --power-min 500 --output high_power.json
        pv-pan-tool export --format csv --efficiency-min 22 --include-metadata --output efficient.csv
        pv-pan-tool export --format csv --height-min 2000 --width-max 1200 --output compact.csv
        pv-pan-tool export --format xlsx --width-min 1000 --height-max 1800 --output medium_size.xlsx
    """
    config = ctx.obj.get('config', {})
    verbose = ctx.obj.get('verbose', False)

    # Get database path
    db_path = get_config('database_path', 'data/database/pv_modules.db',
                        config_file=ctx.obj.get('config_file'))

    # Validate output file extension
    if not validate_output_format(output, export_format):
        console.print(f"[red]Error: Output file extension doesn't match format {export_format}[/red]")
        raise click.Abort()

    try:
        db = PVModuleDatabase(str(db_path))

        # Build search criteria
        criteria = {}
        if manufacturer:
            criteria['manufacturer'] = manufacturer
        if model:
            criteria['model'] = model
        if power_min is not None:
            criteria['power_min'] = power_min
        if power_max is not None:
            criteria['power_max'] = power_max
        if efficiency_min is not None:
            criteria['efficiency_min'] = efficiency_min
        if efficiency_max is not None:
            criteria['efficiency_max'] = efficiency_max
        if cell_type:
            criteria['cell_type'] = cell_type
        if module_type:
            criteria['module_type'] = module_type
        if height_min is not None:
            criteria['height_min'] = height_min
        if height_max is not None:
            criteria['height_max'] = height_max
        if width_min is not None:
            criteria['width_min'] = width_min
        if width_max is not None:
            criteria['width_max'] = width_max

        if verbose:
            console.print(f"[blue]Export criteria:[/blue] {criteria}")
            console.print(f"[blue]Format:[/blue] {export_format}")
            console.print(f"[blue]Output:[/blue] {output}")

        # Get data to export
        with console.status("[bold green]Querying database..."):
            modules = db.search_modules(
                manufacturer=manufacturer,
                model=model,
                min_power=power_min,
                max_power=power_max,
                min_efficiency=efficiency_min,
                max_efficiency=efficiency_max,
                cell_type=cell_type,
                min_height=height_min,
                max_height=height_max,
                min_width=width_min,
                max_width=width_max,
                limit=limit
            )

        if not modules:
            console.print("[yellow]No modules found matching the criteria.[/yellow]")
            return

        console.print(f"[green]Found {len(modules)} modules to export[/green]")

        # Add additional data if requested
        if include_metadata or include_raw:
            modules = enrich_module_data(db, modules, include_metadata, include_raw, verbose)

        # Export data
        export_data(modules, output, export_format, verbose)

        # Show export summary
        show_export_summary(modules, output, export_format)

    except Exception as e:
        console.print(f"[red]Error during export: {e}[/red]")
        if verbose:
            console.print_exception()
        raise click.Abort()


def validate_output_format(output_path, export_format):
    """Validate that output file extension matches export format."""
    extension = output_path.suffix.lower()

    format_extensions = {
        'csv': ['.csv'],
        'xlsx': ['.xlsx', '.xls'],
        'json': ['.json']
    }

    return extension in format_extensions.get(export_format, [])


def enrich_module_data(db, modules, include_metadata, include_raw, verbose):
    """Add additional data to modules if requested."""
    if not (include_metadata or include_raw):
        return modules

    enriched_modules = []

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console
    ) as progress:

        task = progress.add_task("Enriching data...", total=len(modules))

        for module in modules:
            enriched_module = module.copy()

            if include_raw:
                raw_data = db.get_raw_pan_data(module.get('id'))
                if raw_data:
                    enriched_module['raw_pan_data'] = raw_data

            if include_metadata:
                # Metadata is already included in the basic module data
                pass

            enriched_modules.append(enriched_module)
            progress.update(task, advance=1)

    return enriched_modules


def export_data(modules, output_path, export_format, verbose):
    """Export data to the specified format."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if export_format == 'csv':
        export_to_csv(modules, output_path, verbose)
    elif export_format == 'xlsx':
        export_to_excel(modules, output_path, verbose)
    elif export_format == 'json':
        export_to_json(modules, output_path, verbose)


def export_to_csv(modules, output_path, verbose):
    """Export modules to CSV format."""
    try:
        import csv

        if not modules:
            return

        # Get all unique keys
        all_keys = set()
        for module in modules:
            all_keys.update(module.keys())

        sorted_keys = sorted(all_keys)

        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=sorted_keys)
            writer.writeheader()

            for module in modules:
                # Clean None values and ensure all values are strings
                clean_module = {}
                for key in sorted_keys:
                    value = module.get(key)
                    if value is not None:
                        clean_module[key] = str(value)
                    else:
                        clean_module[key] = ''

                writer.writerow(clean_module)

        if verbose:
            console.print(f"[green]CSV export completed: {len(modules)} records[/green]")

    except Exception as e:
        console.print(f"[red]Error exporting to CSV: {e}[/red]")
        raise


def export_to_excel(modules, output_path, verbose):
    """Export modules to Excel format."""
    try:
        import pandas as pd

        if not modules:
            # Create empty Excel file
            pd.DataFrame().to_excel(output_path, index=False)
            return

        # Convert to DataFrame
        df = pd.DataFrame(modules)

        # Clean data
        df = df.fillna('')

        # Create Excel writer with multiple sheets if needed
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='PV_Modules', index=False)

            # Summary sheet
            create_summary_sheet(df, writer)

            # Format sheets
            format_excel_sheets(writer, df)

        if verbose:
            console.print(f"[green]Excel export completed: {len(modules)} records[/green]")

    except ImportError:
        console.print("[red]Error: pandas and openpyxl are required for Excel export[/red]")
        console.print("Install with: pip install pandas openpyxl")
        raise click.Abort()
    except Exception as e:
        console.print(f"[red]Error exporting to Excel: {e}[/red]")
        raise


def export_to_json(modules, output_path, verbose):
    """Export modules to JSON format."""
    try:
        json_content = format_json(modules, indent=2)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(json_content)

        if verbose:
            console.print(f"[green]JSON export completed: {len(modules)} records[/green]")

    except Exception as e:
        console.print(f"[red]Error exporting to JSON: {e}[/red]")
        raise


def create_summary_sheet(df, writer):
    """Create a summary sheet in Excel export."""
    try:
        import pandas as pd

        summary_data = []

        # Basic statistics
        summary_data.append(['Total Modules', len(df)])
        summary_data.append(['Unique Manufacturers', df['manufacturer'].nunique() if 'manufacturer' in df.columns else 0])
        summary_data.append(['Unique Models', df['model'].nunique() if 'model' in df.columns else 0])

        # Power statistics
        if 'pmax_stc' in df.columns:
            power_col = pd.to_numeric(df['pmax_stc'], errors='coerce')
            summary_data.append(['Min Power (W)', power_col.min()])
            summary_data.append(['Max Power (W)', power_col.max()])
            summary_data.append(['Avg Power (W)', power_col.mean()])

        # Efficiency statistics
        if 'efficiency_stc' in df.columns:
            eff_col = pd.to_numeric(df['efficiency_stc'], errors='coerce')
            summary_data.append(['Min Efficiency (%)', eff_col.min()])
            summary_data.append(['Max Efficiency (%)', eff_col.max()])
            summary_data.append(['Avg Efficiency (%)', eff_col.mean()])

        # Create summary DataFrame
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    except Exception as e:
        console.print(f"[yellow]Warning: Could not create summary sheet: {e}[/yellow]")


def format_excel_sheets(writer, df):
    """Format Excel sheets for better readability."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = writer.book

        # Format main sheet
        if 'PV_Modules' in workbook.sheetnames:
            worksheet = workbook['PV_Modules']

            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Format summary sheet
        if 'Summary' in workbook.sheetnames:
            summary_sheet = workbook['Summary']

            for cell in summary_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 15

    except Exception as e:
        console.print(f"[yellow]Warning: Could not format Excel sheets: {e}[/yellow]")


def show_export_summary(modules, output_path, export_format):
    """Show export completion summary."""
    file_size = output_path.stat().st_size if output_path.exists() else 0

    summary_text = f"""
Export completed successfully!

Records exported: {len(modules):,}
Output file: {output_path}
File size: {format_file_size(file_size)}
Format: {export_format.upper()}
"""

    panel = Panel(
        summary_text.strip(),
        title="Export Summary",
        border_style="green"
    )
    console.print(panel)


def format_file_size(size_bytes):
    """Format file size in human-readable format."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.1f} GB"


@click.command()
@click.option(
    "--template", "-t",
    type=click.Choice(['basic', 'detailed', 'comparison', 'statistics']),
    default='basic',
    help="Report template to use"
)
@click.option(
    "--output", "-o",
    type=click.Path(path_type=Path),
    required=True,
    help="Output file path (HTML or PDF)"
)
@click.option(
    "--manufacturer", "-m",
    help="Filter by manufacturer for focused report"
)
@click.option(
    "--include-charts", "-c",
    is_flag=True,
    help="Include charts and visualizations"
)
@click.pass_context
def report(ctx, template, output, manufacturer, include_charts):
    """
    Generate comprehensive reports.

    Create detailed HTML or PDF reports with statistics,
    comparisons, and visualizations.

    Examples:
        pv-pan-tool report --template detailed --output report.html
        pv-pan-tool report --template comparison --manufacturer "Jinko" --output jinko_report.pdf
        pv-pan-tool report --template statistics --include-charts --output stats.html
    """
    config = ctx.obj.get('config', {})
    verbose = ctx.obj.get('verbose', False)

    console.print("[yellow]Report generation feature coming soon![/yellow]")
    console.print("This will generate comprehensive HTML/PDF reports with:")
    console.print("• Statistical analysis")
    console.print("• Module comparisons")
    console.print("• Charts and visualizations")
    console.print("• Manufacturer profiles")


# Remove the problematic line at the end
# export.add_command(report, name="report")

    console.print("• Charts and visualizations")
    console.print("• Manufacturer profiles")


# Remove the problematic line at the end
# export.add_command(report, name="report")
