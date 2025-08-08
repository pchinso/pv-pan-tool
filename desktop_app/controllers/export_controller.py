"""
Export controller for PV PAN Tool Desktop Application.

This module handles data export operations in various formats.
"""

import json
import csv
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from .database_controller import DatabaseController


class ExportController:
    """Controller for export operations."""
    
    def __init__(self, db_controller: DatabaseController):
        """
        Initialize the export controller.
        
        Args:
            db_controller: Database controller instance
        """
        self.db_controller = db_controller
        self.supported_formats = ["csv", "json", "xlsx"]
    
    def export_modules(self, modules: List[Dict[str, Any]], 
                      file_path: str, format: str = "csv",
                      include_metadata: bool = False) -> Dict[str, Any]:
        """
        Export modules to file.
        
        Args:
            modules: List of modules to export
            file_path: Output file path
            format: Export format (csv, json, xlsx)
            include_metadata: Include additional metadata
            
        Returns:
            Dictionary with export results
        """
        try:
            if format not in self.supported_formats:
                return {
                    "success": False,
                    "error": f"Unsupported format: {format}. Supported: {self.supported_formats}"
                }
            
            if not modules:
                return {
                    "success": False,
                    "error": "No modules to export"
                }
            
            # Prepare data for export
            export_data = self._prepare_export_data(modules, include_metadata)
            
            # Export based on format
            if format == "csv":
                result = self._export_csv(export_data, file_path)
            elif format == "json":
                result = self._export_json(export_data, file_path, include_metadata)
            elif format == "xlsx":
                result = self._export_xlsx(export_data, file_path, include_metadata)
            else:
                return {"success": False, "error": f"Format {format} not implemented"}
            
            if result["success"]:
                result.update({
                    "exported_count": len(modules),
                    "file_path": file_path,
                    "format": format,
                    "timestamp": datetime.now().isoformat()
                })
            
            return result
            
        except Exception as e:
            return {
                "success": False,
                "error": f"Export failed: {str(e)}"
            }
    
    def _prepare_export_data(self, modules: List[Dict[str, Any]], 
                           include_metadata: bool) -> List[Dict[str, Any]]:
        """
        Prepare module data for export.
        
        Args:
            modules: Raw module data
            include_metadata: Include metadata fields
            
        Returns:
            Prepared data for export
        """
        export_data = []
        
        for module in modules:
            # Create a clean copy
            clean_module = {}
            
            # Basic module information
            basic_fields = [
                "id", "manufacturer", "model", "series",
                "pmax_stc", "vmp_stc", "imp_stc", "voc_stc", "isc_stc",
                "efficiency_stc", "cell_type", "module_type",
                "length", "width", "thickness", "weight",
                "cells_in_series", "total_cells"
            ]
            
            for field in basic_fields:
                if field in module:
                    clean_module[field] = module[field]
            
            # Temperature coefficients
            temp_coeff_fields = [
                "temp_coeff_pmax", "temp_coeff_voc", "temp_coeff_isc"
            ]
            
            for field in temp_coeff_fields:
                if field in module:
                    clean_module[field] = module[field]
            
            # Additional electrical parameters
            additional_fields = [
                "vmp_noct", "imp_noct", "pmax_noct",
                "noct", "operating_temp_min", "operating_temp_max"
            ]
            
            for field in additional_fields:
                if field in module:
                    clean_module[field] = module[field]
            
            # Include metadata if requested
            if include_metadata:
                metadata_fields = [
                    "file_path", "file_name", "file_size",
                    "parsed_at", "unique_id"
                ]
                
                for field in metadata_fields:
                    if field in module:
                        clean_module[field] = module[field]
            
            export_data.append(clean_module)
        
        return export_data
    
    def _export_csv(self, data: List[Dict[str, Any]], file_path: str) -> Dict[str, Any]:
        """
        Export data to CSV format.
        
        Args:
            data: Data to export
            file_path: Output file path
            
        Returns:
            Export result
        """
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                if data:
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for row in data:
                        # Clean None values
                        clean_row = {k: (v if v is not None else '') for k, v in row.items()}
                        writer.writerow(clean_row)
            
            return {"success": True}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _export_json(self, data: List[Dict[str, Any]], file_path: str,
                    include_metadata: bool) -> Dict[str, Any]:
        """
        Export data to JSON format.
        
        Args:
            data: Data to export
            file_path: Output file path
            include_metadata: Include export metadata
            
        Returns:
            Export result
        """
        try:
            export_object = {
                "export_info": {
                    "timestamp": datetime.now().isoformat(),
                    "total_modules": len(data),
                    "format": "json",
                    "version": "1.0"
                },
                "modules": data
            }
            
            if not include_metadata:
                export_object = data
            
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_object, jsonfile, indent=2, default=str, ensure_ascii=False)
            
            return {"success": True}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _export_xlsx(self, data: List[Dict[str, Any]], file_path: str,
                    include_metadata: bool) -> Dict[str, Any]:
        """
        Export data to Excel format.
        
        Args:
            data: Data to export
            file_path: Output file path
            include_metadata: Include metadata sheets
            
        Returns:
            Export result
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Main data sheet
            ws_main = wb.create_sheet("PV Modules")
            
            # Add data to sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_main.append(r)
            
            # Format header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws_main[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws_main.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws_main.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet if metadata requested
            if include_metadata:
                ws_summary = wb.create_sheet("Summary")
                
                # Calculate summary statistics
                summary_data = self._calculate_summary_stats(data)
                
                # Add summary data
                ws_summary.append(["Metric", "Value"])
                for metric, value in summary_data.items():
                    ws_summary.append([metric, value])
                
                # Format summary sheet
                for cell in ws_summary[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                ws_summary.column_dimensions['A'].width = 25
                ws_summary.column_dimensions['B'].width = 15
            
            # Save workbook
            wb.save(file_path)
            
            return {"success": True}
            
        except ImportError:
            return {
                "success": False,
                "error": "Excel export requires pandas and openpyxl. Install with: pip install pandas openpyxl"
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _calculate_summary_stats(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Calculate summary statistics for export.
        
        Args:
            data: Module data
            
        Returns:
            Summary statistics
        """
        if not data:
            return {}
        
        summary = {
            "Total Modules": len(data),
            "Export Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Calculate power statistics
        powers = [m.get("pmax_stc") for m in data if m.get("pmax_stc") is not None]
        if powers:
            summary.update({
                "Min Power (W)": min(powers),
                "Max Power (W)": max(powers),
                "Avg Power (W)": round(sum(powers) / len(powers), 1)
            })
        
        # Calculate efficiency statistics
        efficiencies = [m.get("efficiency_stc") for m in data if m.get("efficiency_stc") is not None]
        if efficiencies:
            summary.update({
                "Min Efficiency (%)": round(min(efficiencies), 2),
                "Max Efficiency (%)": round(max(efficiencies), 2),
                "Avg Efficiency (%)": round(sum(efficiencies) / len(efficiencies), 2)
            })
        
        # Count manufacturers
        manufacturers = set(m.get("manufacturer") for m in data if m.get("manufacturer"))
        summary["Unique Manufacturers"] = len(manufacturers)
        
        # Count cell types
        cell_types = set(m.get("cell_type") for m in data if m.get("cell_type"))
        summary["Unique Cell Types"] = len(cell_types)
        
        return summary
    
    def export_search_results(self, search_results: Dict[str, Any],
                            file_path: str, format: str = "csv") -> Dict[str, Any]:
        """
        Export search results with search metadata.
        
        Args:
            search_results: Search results from SearchController
            file_path: Output file path
            format: Export format
            
        Returns:
            Export result
        """
        try:
            modules = search_results.get("modules", [])
            
            if format == "json":
                # Include search metadata in JSON export
                export_data = {
                    "search_info": {
                        "timestamp": datetime.now().isoformat(),
                        "search_params": search_results.get("search_params", {}),
                        "total_results": search_results.get("total_results", 0)
                    },
                    "modules": modules
                }
                
                with open(file_path, 'w', encoding='utf-8') as jsonfile:
                    json.dump(export_data, jsonfile, indent=2, default=str, ensure_ascii=False)
                
                return {
                    "success": True,
                    "exported_count": len(modules),
                    "file_path": file_path
                }
            else:
                # For other formats, just export the modules
                return self.export_modules(modules, file_path, format, include_metadata=True)
                
        except Exception as e:
            return {
                "success": False,
                "error": f"Export failed: {str(e)}"
            }
    
    def export_comparison(self, comparison_data: Dict[str, Any],
                         file_path: str, format: str = "csv") -> Dict[str, Any]:
        """
        Export module comparison data.
        
        Args:
            comparison_data: Comparison data from CompareWidget
            file_path: Output file path
            format: Export format
            
        Returns:
            Export result
        """
        try:
            modules = comparison_data.get("modules", [])
            
            if format == "json":
                # Include comparison metadata
                export_data = {
                    "comparison_info": {
                        "timestamp": datetime.now().isoformat(),
                        "compared_modules": len(modules),
                        "comparison_type": comparison_data.get("type", "manual")
                    },
                    "modules": modules,
                    "analysis": comparison_data.get("analysis", {})
                }
                
                with open(file_path, 'w', encoding='utf-8') as jsonfile:
                    json.dump(export_data, jsonfile, indent=2, default=str, ensure_ascii=False)
                
                return {
                    "success": True,
                    "exported_count": len(modules),
                    "file_path": file_path
                }
            else:
                return self.export_modules(modules, file_path, format, include_metadata=True)
                
        except Exception as e:
            return {
                "success": False,
                "error": f"Export failed: {str(e)}"
            }
    
    def get_export_formats(self) -> List[Dict[str, str]]:
        """
        Get available export formats.
        
        Returns:
            List of format information
        """
        return [
            {
                "format": "csv",
                "name": "CSV (Comma Separated Values)",
                "extension": ".csv",
                "description": "Compatible with Excel and other spreadsheet applications"
            },
            {
                "format": "json",
                "name": "JSON (JavaScript Object Notation)",
                "extension": ".json",
                "description": "Structured data format for programming and APIs"
            },
            {
                "format": "xlsx",
                "name": "Excel Workbook",
                "extension": ".xlsx",
                "description": "Microsoft Excel format with multiple sheets and formatting"
            }
        ]
    
    def validate_export_path(self, file_path: str, format: str) -> Dict[str, Any]:
        """
        Validate export file path.
        
        Args:
            file_path: Proposed file path
            format: Export format
            
        Returns:
            Validation result
        """
        try:
            path = Path(file_path)
            
            # Check if directory exists
            if not path.parent.exists():
                return {
                    "valid": False,
                    "error": f"Directory does not exist: {path.parent}"
                }
            
            # Check if directory is writable
            if not path.parent.is_dir():
                return {
                    "valid": False,
                    "error": f"Parent path is not a directory: {path.parent}"
                }
            
            # Check file extension
            expected_extensions = {
                "csv": ".csv",
                "json": ".json",
                "xlsx": ".xlsx"
            }
            
            expected_ext = expected_extensions.get(format)
            if expected_ext and not file_path.lower().endswith(expected_ext):
                return {
                    "valid": False,
                    "error": f"File extension should be {expected_ext} for {format} format"
                }
            
            # Check if file already exists
            if path.exists():
                return {
                    "valid": True,
                    "warning": f"File already exists and will be overwritten: {file_path}"
                }
            
            return {"valid": True}
            
        except Exception as e:
            return {
                "valid": False,
                "error": f"Invalid file path: {str(e)}"
            }

